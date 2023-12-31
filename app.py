from flask import Flask, render_template, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

class Atividade(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    descricao = db.Column(db.String(255), nullable=False)
    data = db.Column(db.String(20), nullable=False)

# Rotas para a aplicação
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/registrar_opcao', methods=['POST'])
def registrar_opcao():
    data = request.get_json()

    atividade = Atividade(descricao=data['opcoes'][0], data='2023-12-03')  # Substitua '2023-12-03' pela data real
    db.session.add(atividade)
    db.session.commit()

    return jsonify({'message': 'Atividade registrada com sucesso!'})

@app.route('/relatorio', methods=['GET'])
# ... (código anterior permanece inalterado)

@app.route('/relatorio', methods=['GET'])
def relatorio():
    filtro = request.args.get('filtro')

    # Lógica para gerar o relatório com base no filtro
    if filtro in ['dia', 'semana', 'mes', 'ano']:
        # Exemplo simples: obter todas as atividades do banco de dados
        atividades = Atividade.query.all()
        # Geração do XLSX
        generate_xlsx(atividades, filtro)
        return jsonify({'success': True, 'message': f'Relatório gerado com sucesso para filtro: {filtro}'})

    return jsonify({'success': False, 'error': 'Filtro não suportado'})

# ... (código posterior permanece inalterado)


# ... (o código anterior permanece inalterado)

def generate_xlsx(atividades, filtro):
    filename = f"relatorio_{filtro}.xlsx"
    document_title = "Relatório Atividades"

    wb = Workbook()
    ws = wb.active
    ws.title = document_title

    # Adicionar título
    ws['A1'] = document_title
    ws['A2'] = ''

    # Adicionar títulos de coluna
    ws['A3'] = 'Tarefa'
    ws['B3'] = 'Quantidade'
    ws['C3'] = ''  # Linha em branco entre títulos e dados

    # Adicionar dados
    dados_relatorio = {}

    for atividade in atividades:
        if atividade.descricao not in dados_relatorio:
            dados_relatorio[atividade.descricao] = 1
        else:
            dados_relatorio[atividade.descricao] += 1

    # Adicionar nome da tarefa e quantidade de vezes registrada
    for i, (tarefa, quantidade) in enumerate(dados_relatorio.items(), start=4):
        ws[f'A{i}'] = f"{tarefa}"
        ws[f'B{i}'] = f"{quantidade}"
        ws[f'C{i}'] = ''  # Linha em branco entre atividades

    # Ajustar alinhamento
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Adicionar timestamp ao nome do arquivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename_with_timestamp = f"{filename[:-5]}_{timestamp}.xlsx"

    # Salvar o XLSX
    wb.save(filename_with_timestamp)

# ... (o código posterior permanece inalterado)


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run()

